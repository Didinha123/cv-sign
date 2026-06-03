from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random

wb = Workbook()

# ── Paleta ───────────────────────────────────────────────────────────
C_PURPLE     = "6E2C7C"
C_PURPLE_LT  = "EDE0F2"
C_PINK       = "E63946"
C_PINK_LT    = "FFE8EA"
C_GREEN      = "166534"
C_GREEN_LT   = "DCFCE7"
C_YELLOW_LT  = "FFF9C4"
C_BLUE_LT    = "DBEAFE"
C_GRAY       = "F5F5F5"
C_WHITE      = "FFFFFF"
C_TEXT       = "2B2D42"
C_DARK       = "1A1A2E"

def hdr(ws, row, col, text, bg=C_PURPLE, fg=C_WHITE, bold=True, sz=11, wrap=False, center=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, color=fg, size=sz, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center", wrap_text=wrap)
    return c

def cell(ws, row, col, value, bold=False, color=C_TEXT, bg=None, fmt=None,
         align="left", wrap=False, sz=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, size=sz, name="Arial")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c

def border_range(ws, min_row, max_row, min_col, max_col, color="CCCCCC"):
    thin = Side(style="thin", color=color)
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row+1):
        for c_ in range(min_col, max_col+1):
            ws.cell(r, c_).border = b

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def tab_color(ws, color):
    ws.sheet_properties.tabColor = color

def freeze(ws, cell_ref):
    ws.freeze_panes = cell_ref

# ═══════════════════════════════════════════════════════════════════════
# ABA 1 — PEDIDOS
# ═══════════════════════════════════════════════════════════════════════
ws_p = wb.active
ws_p.title = "📋 Pedidos"
tab_color(ws_p, C_PURPLE)
ws_p.row_dimensions[1].height = 36

headers_p = [
    "ID Pedido","Data","Hora","Cliente","Telefone",
    "Rua + Nº","Bairro","Cidade",
    "Itens do Pedido","Qtd","Subtotal (R$)","Taxa Entrega (R$)","Total (R$)",
    "Pagamento","Status","Troco Para","Observações"
]
for i, h in enumerate(headers_p, 1):
    hdr(ws_p, 1, i, h, bg=C_PURPLE, sz=10, wrap=True)

# Dados de exemplo
pedidos_ex = [
    ["#00001","2026-06-01","14:32","Maria Silva","(19) 99111-2222",
     "Rua das Flores, 45","Centro","Mogi Mirim",
     "2x Copo Thaina (500ml), 1x Copo Diego (300ml)",3,73.00,5.00,"=K2+L2",
     "PIX","Entregue","",""],
    ["#00002","2026-06-01","15:10","João Pereira","(19) 98222-3333",
     "Av. Brasil, 120","Jardim Europa","Mogi Mirim",
     "1x Combo Família",1,69.90,5.00,"=K3+L3",
     "Débito","Entregue","","Portão azul"],
    ["#00003","2026-06-02","16:45","Ana Souza","(19) 97333-4444",
     "Rua Ipiranga, 8","Recanto Di Verona","Mogi Mirim",
     "1x Copo Arthur (700ml), 1x Copo TDA (500ml)",2,58.00,5.00,"=K4+L4",
     "Crédito","Em preparo","",""],
    ["#00004","2026-06-02","17:20","Carlos Lima","(19) 96444-5555",
     "Rua XV de Nov., 200","Vila Nova","Mogi Mirim",
     "3x Copo Davi (300ml)",3,63.00,5.00,"=K5+L5",
     "PIX","Saiu para entrega","",""],
    ["#00005","2026-06-03","14:00","Fernanda Costa","(19) 95555-6666",
     "Rua das Palmeiras, 33","Centro","Mogi Mirim",
     "2x Copo Diego (500ml)",2,52.00,5.00,"=K6+L6",
     "Dinheiro","Aguardando","R$ 60,00",""],
]
status_colors = {
    "Entregue":        C_GREEN_LT,
    "Em preparo":      C_YELLOW_LT,
    "Saiu para entrega": C_BLUE_LT,
    "Aguardando":      "FFF3E0",
    "Cancelado":       C_PINK_LT,
}
for i, row in enumerate(pedidos_ex, 2):
    st = row[14]
    bg = status_colors.get(st, C_WHITE)
    for j, val in enumerate(row, 1):
        c = ws_p.cell(row=i, column=j, value=val)
        c.font = Font(name="Arial", size=10, color=C_TEXT)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center" if j in [1,2,3,6,10,11,12,13,14,15] else "left",
                                  vertical="center")
        if j in [11,12,13]:
            c.number_format = 'R$ #,##0.00'

# Validação de status
dv_status = DataValidation(
    type="list",
    formula1='"Aguardando,Em preparo,Saiu para entrega,Entregue,Cancelado"',
    allow_blank=True
)
ws_p.add_data_validation(dv_status)
dv_status.sqref = "O2:O1000"

# Validação pagamento
dv_pag = DataValidation(
    type="list",
    formula1='"PIX,Débito,Crédito,Dinheiro"',
    allow_blank=True
)
ws_p.add_data_validation(dv_pag)
dv_pag.sqref = "N2:N1000"

border_range(ws_p, 1, len(pedidos_ex)+1, 1, len(headers_p))
freeze(ws_p, "D2")
set_col_widths(ws_p, {
    1:10, 2:12, 3:8, 4:18, 5:16, 6:22, 7:16, 8:14,
    9:38, 10:6, 11:13, 12:14, 13:12,
    14:12, 15:18, 16:12, 17:20
})
ws_p.row_dimensions[1].height = 40
for r in range(2, len(pedidos_ex)+2):
    ws_p.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════
# ABA 2 — CLIENTES
# ═══════════════════════════════════════════════════════════════════════
ws_c = wb.create_sheet("👥 Clientes")
tab_color(ws_c, "4A90A4")

headers_c = [
    "Nome","Telefone","Rua","Número","Bairro","Complemento","Cidade",
    "Total Pedidos","Valor Total Gasto (R$)","Ticket Médio (R$)",
    "Primeiro Pedido","Último Pedido","Cliente Fiel?"
]
for i, h in enumerate(headers_c, 1):
    hdr(ws_c, 1, i, h, bg="4A90A4", sz=10, wrap=True)

clientes_ex = [
    ["Maria Silva","(19) 99111-2222","Rua das Flores","45","Centro","","Mogi Mirim",
     '=COUNTIF(\'📋 Pedidos\'!D:D,A2)',
     '=SUMIF(\'📋 Pedidos\'!D:D,A2,\'📋 Pedidos\'!M:M)',
     '=IFERROR(I2/H2,0)',
     "2026-06-01","2026-06-01",'=IF(H2>=10,"🌟 Sim","Não")'],
    ["João Pereira","(19) 98222-3333","Av. Brasil","120","Jardim Europa","","Mogi Mirim",
     '=COUNTIF(\'📋 Pedidos\'!D:D,A3)',
     '=SUMIF(\'📋 Pedidos\'!D:D,A3,\'📋 Pedidos\'!M:M)',
     '=IFERROR(I3/H3,0)',
     "2026-06-01","2026-06-01",'=IF(H3>=10,"🌟 Sim","Não")'],
    ["Ana Souza","(19) 97333-4444","Rua Ipiranga","8","Recanto Di Verona","Apto 3","Mogi Mirim",
     '=COUNTIF(\'📋 Pedidos\'!D:D,A4)',
     '=SUMIF(\'📋 Pedidos\'!D:D,A4,\'📋 Pedidos\'!M:M)',
     '=IFERROR(I4/H4,0)',
     "2026-06-02","2026-06-02",'=IF(H4>=10,"🌟 Sim","Não")'],
    ["Carlos Lima","(19) 96444-5555","Rua XV de Nov.","200","Vila Nova","","Mogi Mirim",
     '=COUNTIF(\'📋 Pedidos\'!D:D,A5)',
     '=SUMIF(\'📋 Pedidos\'!D:D,A5,\'📋 Pedidos\'!M:M)',
     '=IFERROR(I5/H5,0)',
     "2026-06-02","2026-06-02",'=IF(H5>=10,"🌟 Sim","Não")'],
    ["Fernanda Costa","(19) 95555-6666","Rua das Palmeiras","33","Centro","","Mogi Mirim",
     '=COUNTIF(\'📋 Pedidos\'!D:D,A6)',
     '=SUMIF(\'📋 Pedidos\'!D:D,A6,\'📋 Pedidos\'!M:M)',
     '=IFERROR(I6/H6,0)',
     "2026-06-03","2026-06-03",'=IF(H6>=10,"🌟 Sim","Não")'],
]
for i, row in enumerate(clientes_ex, 2):
    bg = C_GRAY if i % 2 == 0 else C_WHITE
    for j, val in enumerate(row, 1):
        c = ws_c.cell(row=i, column=j, value=val)
        c.font = Font(name="Arial", size=10, color=C_TEXT)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center" if j in [8,9,10,13] else "left",
                                  vertical="center")
        if j in [9, 10]:
            c.number_format = 'R$ #,##0.00'

border_range(ws_c, 1, len(clientes_ex)+1, 1, len(headers_c))
freeze(ws_c, "B2")
set_col_widths(ws_c, {
    1:20, 2:18, 3:22, 4:8, 5:18, 6:16, 7:14,
    8:13, 9:20, 10:16, 11:14, 12:14, 13:14
})
ws_c.row_dimensions[1].height = 40
for r in range(2, len(clientes_ex)+2):
    ws_c.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════
# ABA 3 — CARDÁPIO
# ═══════════════════════════════════════════════════════════════════════
ws_m = wb.create_sheet("🛍️ Cardápio")
tab_color(ws_m, C_PINK)

headers_m = ["Produto","Tamanho","Preço (R$)","Categoria","Disponível","Descrição"]
for i, h in enumerate(headers_m, 1):
    hdr(ws_m, 1, i, h, bg=C_PINK, sz=10)

menu_items = [
    ["Copo Diego",  "300ml", 21.00, "Copo Individual","✅ Sim","Açaí cremoso, leite condensado, paçoca, amendoim"],
    ["Copo Diego",  "500ml", 26.00, "Copo Individual","✅ Sim","Açaí cremoso, leite condensado, paçoca, amendoim"],
    ["Copo Diego",  "700ml", 30.00, "Copo Individual","✅ Sim","Açaí cremoso, leite condensado, paçoca, amendoim"],
    ["Copo Arthur", "300ml", 22.00, "Copo Individual","✅ Sim","Morango, creme de avelã, leite Ninho"],
    ["Copo Arthur", "500ml", 27.00, "Copo Individual","✅ Sim","Morango, creme de avelã, leite Ninho"],
    ["Copo Arthur", "700ml", 31.00, "Copo Individual","✅ Sim","Morango, creme de avelã, leite Ninho"],
    ["Copo Thaina", "300ml", 21.00, "Copo Individual","✅ Sim","Morango, leite condensado, amendoim crocante"],
    ["Copo Thaina", "500ml", 26.00, "Copo Individual","✅ Sim","Morango, leite condensado, amendoim crocante"],
    ["Copo Thaina", "700ml", 30.00, "Copo Individual","✅ Sim","Morango, leite condensado, amendoim crocante"],
    ["Copo Davi",   "300ml", 21.00, "Copo Individual","✅ Sim","Banana, leite condensado, granola crocante"],
    ["Copo Davi",   "500ml", 26.00, "Copo Individual","✅ Sim","Banana, leite condensado, granola crocante"],
    ["Copo Davi",   "700ml", 30.00, "Copo Individual","✅ Sim","Banana, leite condensado, granola crocante"],
    ["Copo TDA",    "300ml", 22.00, "Copo Individual","✅ Sim","Uva, banana, leite em pó, creme de avelã"],
    ["Copo TDA",    "500ml", 27.00, "Copo Individual","✅ Sim","Uva, banana, leite em pó, creme de avelã"],
    ["Copo TDA",    "700ml", 31.00, "Copo Individual","✅ Sim","Uva, banana, leite em pó, creme de avelã"],
    ["Combo Família","Único",69.90, "Combo",          "✅ Sim","4 copos, acompanhamentos sortidos"],
]
cat_colors = {
    "Copo Individual": C_PURPLE_LT,
    "Combo":           C_PINK_LT,
}
for i, row in enumerate(menu_items, 2):
    bg = cat_colors.get(row[3], C_WHITE)
    for j, val in enumerate(row, 1):
        c = ws_m.cell(row=i, column=j, value=val)
        c.font = Font(name="Arial", size=10,
                       bold=(j==1), color=C_PURPLE if j==1 else C_TEXT)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center" if j in [2,3,4,5] else "left",
                                  vertical="center")
        if j == 3:
            c.number_format = 'R$ #,##0.00'

border_range(ws_m, 1, len(menu_items)+1, 1, len(headers_m))
set_col_widths(ws_m, {1:16, 2:10, 3:14, 4:18, 5:12, 6:40})
ws_m.row_dimensions[1].height = 36
for r in range(2, len(menu_items)+2):
    ws_m.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════
# ABA 4 — FINANCEIRO
# ═══════════════════════════════════════════════════════════════════════
ws_f = wb.create_sheet("💰 Financeiro")
tab_color(ws_f, C_GREEN)

# Título
ws_f.merge_cells("A1:F1")
c = ws_f["A1"]
c.value = "💰 Resumo Financeiro — TDA Açaí"
c.font = Font(bold=True, size=14, color=C_WHITE, name="Arial")
c.fill = PatternFill("solid", fgColor=C_GREEN)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_f.row_dimensions[1].height = 36

# KPIs
kpi_headers = ["📊 Indicador","Valor"]
for i, h in enumerate(kpi_headers, 1):
    hdr(ws_f, 3, i, h, bg="166534", sz=11)
    hdr(ws_f, 3, i+2, h, bg="166534", sz=11)

kpis_left = [
    ["Total de Pedidos",        "=COUNTA('📋 Pedidos'!A2:A10000)-COUNTBLANK('📋 Pedidos'!A2:A10000)"],
    ["Pedidos Entregues",       "=COUNTIF('📋 Pedidos'!O:O,\"Entregue\")"],
    ["Pedidos em Andamento",    "=COUNTIFS('📋 Pedidos'!O:O,\"<>Entregue\",'📋 Pedidos'!O:O,\"<>Cancelado\",'📋 Pedidos'!O:O,\"<>\")"],
    ["Pedidos Cancelados",      "=COUNTIF('📋 Pedidos'!O:O,\"Cancelado\")"],
    ["Total de Clientes",       "=COUNTA('👥 Clientes'!A2:A10000)-COUNTBLANK('👥 Clientes'!A2:A10000)"],
    ["Ticket Médio (R$)",       "=IFERROR(AVERAGE('📋 Pedidos'!M2:M10000),0)"],
]
kpis_right = [
    ["Faturamento Total (R$)",  "=SUMIF('📋 Pedidos'!O:O,\"Entregue\",'📋 Pedidos'!M:M)"],
    ["Receita com Entrega (R$)","=SUMIF('📋 Pedidos'!O:O,\"Entregue\",'📋 Pedidos'!L:L)"],
    ["Receita Produtos (R$)",   "=SUMIF('📋 Pedidos'!O:O,\"Entregue\",'📋 Pedidos'!K:K)"],
    ["Pagamentos PIX (R$)",     "=SUMIFS('📋 Pedidos'!M:M,'📋 Pedidos'!N:N,\"PIX\",'📋 Pedidos'!O:O,\"Entregue\")"],
    ["Pagamentos Cartão (R$)",  "=SUMIFS('📋 Pedidos'!M:M,'📋 Pedidos'!N:N,\"Débito\",'📋 Pedidos'!O:O,\"Entregue\")+SUMIFS('📋 Pedidos'!M:M,'📋 Pedidos'!N:N,\"Crédito\",'📋 Pedidos'!O:O,\"Entregue\")"],
    ["Pagamentos Dinheiro (R$)","=SUMIFS('📋 Pedidos'!M:M,'📋 Pedidos'!N:N,\"Dinheiro\",'📋 Pedidos'!O:O,\"Entregue\")"],
]
money_rows_r = {4,5,6,7,8,9}  # linhas com R$

for i, (label, formula) in enumerate(kpis_left, 4):
    cell(ws_f, i, 1, label, bold=True, bg=C_GRAY, sz=10)
    c2 = ws_f.cell(row=i, column=2, value=formula)
    c2.font = Font(name="Arial", size=10, bold=True, color=C_GREEN)
    c2.fill = PatternFill("solid", fgColor=C_GREEN_LT)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    if "R$" in label:
        c2.number_format = 'R$ #,##0.00'

for i, (label, formula) in enumerate(kpis_right, 4):
    cell(ws_f, i, 3, label, bold=True, bg=C_GRAY, sz=10)
    c2 = ws_f.cell(row=i, column=4, value=formula)
    c2.font = Font(name="Arial", size=10, bold=True, color=C_GREEN)
    c2.fill = PatternFill("solid", fgColor=C_GREEN_LT)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    if "R$" in label:
        c2.number_format = 'R$ #,##0.00'

# Tabela por forma de pagamento
hdr(ws_f, 11, 1, "Forma de Pagamento", bg=C_PURPLE, sz=10)
hdr(ws_f, 11, 2, "Qtd Pedidos",        bg=C_PURPLE, sz=10)
hdr(ws_f, 11, 3, "Valor Total (R$)",   bg=C_PURPLE, sz=10)
hdr(ws_f, 11, 4, "% do Total",         bg=C_PURPLE, sz=10)

payments = ["PIX","Débito","Crédito","Dinheiro"]
for i, p in enumerate(payments, 12):
    cell(ws_f, i, 1, p, bg=C_WHITE, sz=10)
    c_q = ws_f.cell(row=i, column=2,
        value=f'=COUNTIFS(\'📋 Pedidos\'!N:N,"{p}",\'📋 Pedidos\'!O:O,"Entregue")')
    c_q.font = Font(name="Arial", size=10); c_q.alignment = Alignment(horizontal="center",vertical="center")
    c_v = ws_f.cell(row=i, column=3,
        value=f'=SUMIFS(\'📋 Pedidos\'!M:M,\'📋 Pedidos\'!N:N,"{p}",\'📋 Pedidos\'!O:O,"Entregue")')
    c_v.font = Font(name="Arial", size=10); c_v.number_format = 'R$ #,##0.00'
    c_v.alignment = Alignment(horizontal="center",vertical="center")
    c_pct = ws_f.cell(row=i, column=4,
        value=f'=IFERROR(C{i}/SUM($C$12:$C$15),0)')
    c_pct.font = Font(name="Arial", size=10); c_pct.number_format = '0.0%'
    c_pct.alignment = Alignment(horizontal="center",vertical="center")

# Tabela por produto
hdr(ws_f, 11, 6, "Produto + Tamanho",  bg=C_PINK, sz=10)
hdr(ws_f, 11, 7, "Preço Unit. (R$)",   bg=C_PINK, sz=10)
hdr(ws_f, 11, 8, "Qtd Vendida",        bg=C_PINK, sz=10)

border_range(ws_f, 3, 16, 1, 8)
set_col_widths(ws_f, {1:26, 2:16, 3:24, 4:16, 5:4, 6:22, 7:16, 8:14})
for r in [1,3,11]:
    ws_f.row_dimensions[r].height = 32
for r in range(4, 17):
    ws_f.row_dimensions[r].height = 22

# ═══════════════════════════════════════════════════════════════════════
# ABA 5 — CONFIG
# ═══════════════════════════════════════════════════════════════════════
ws_cfg = wb.create_sheet("⚙️ Config")
tab_color(ws_cfg, "888888")

ws_cfg.merge_cells("A1:C1")
c = ws_cfg["A1"]
c.value = "⚙️ Configurações do Sistema — TDA Açaí"
c.font = Font(bold=True, size=13, color=C_WHITE, name="Arial")
c.fill = PatternFill("solid", fgColor="444444")
c.alignment = Alignment(horizontal="center", vertical="center")
ws_cfg.row_dimensions[1].height = 34

configs = [
    ("🚚 Taxa de Entrega Padrão (R$)", 5.00,      "Altere aqui para refletir no sistema"),
    ("📲 Chave PIX",                   "",         "CPF, e-mail, telefone ou chave aleatória"),
    ("👤 Nome Recebedor PIX",          "TDA Acai", "Máx. 25 caracteres"),
    ("🏙️ Cidade PIX",                  "Mogi Mirim","Máx. 15 caracteres"),
    ("📱 WhatsApp",                    "(19) 99419-4916", "Número para pedidos"),
    ("🎁 Pedidos p/ Brinde (fidelidade)", 10,     "Qtd de pedidos para ganhar 1 grátis"),
    ("🔐 Senha Admin",                 "1234",     "Altere após primeiro acesso"),
]
hdr(ws_cfg, 2, 1, "Configuração",  bg="555555", sz=10)
hdr(ws_cfg, 2, 2, "Valor",         bg="555555", sz=10)
hdr(ws_cfg, 2, 3, "Observação",    bg="555555", sz=10)

for i, (k, v, obs) in enumerate(configs, 3):
    cell(ws_cfg, i, 1, k,   bold=True, bg=C_GRAY, sz=10)
    c2 = ws_cfg.cell(row=i, column=2, value=v)
    c2.font = Font(name="Arial", size=10, bold=True, color="000080")
    c2.fill = PatternFill("solid", fgColor=C_YELLOW_LT)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    if isinstance(v, float):
        c2.number_format = 'R$ #,##0.00'
    cell(ws_cfg, i, 3, obs, bg=C_WHITE, color="666666", sz=9)

border_range(ws_cfg, 1, len(configs)+2, 1, 3)
set_col_widths(ws_cfg, {1:34, 2:26, 3:38})
for r in range(1, len(configs)+3):
    ws_cfg.row_dimensions[r].height = 26

# ── Salvar ────────────────────────────────────────────────────────────
out = "/sessions/hopeful-dreamy-ritchie/mnt/outputs/TDA_Acai_BancoDeDados.xlsx"
wb.save(out)
print("OK:", out)
